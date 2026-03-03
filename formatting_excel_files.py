
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Specify the directory containing Excel files
directory = "./entreprises"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

# Loop through all files in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        # Load the Excel file
        excel_path = os.path.join(directory, filename)
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Apply formatting to header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Apply borders and wrap text to all cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Auto-adjust column widths to fix overflow
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width to avoid excessive size
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the formatted workbook
        wb.save(excel_path)
        print(f"Formatted {filename}")